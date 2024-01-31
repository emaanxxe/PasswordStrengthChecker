from tkinter import * 
from PIL import ImageTk, Image
import openpyxl
from openpyxl import Workbook
import re


fenetre = Tk()   #Tk() est une class dans tkinter ,fenetre est une variable
fenetre.geometry('500x550+300+150')
fenetre.resizable(False,False)
fenetre.title('Password Strength Checker')#F8F2ED
fenetre.config(background='#002366') #pour determiner la color de background de la fenetre
fenetre.iconbitmap('C:\\Users\\pc\\Desktop\\Python\\icon.ico')
fenetre.attributes('-alpha',0.8)

#fonction pour le champ vide 
def valider(enent=None):
    if mot_de_passe.get() == "":
        erreur="Le champ est obligatoire !"

#fonction Reset:
def Reset(enent=None):
    #En1.delete('0',END)
    #En2.delete('0',END)
    mot_de_passe.delete('0',END)

#la fonction associer a la button Tester :
def check_password_strength(password):
    
    # Définir les critères de force du mot de passe
    length_regex = re.compile(r'.{8,}')
    uppercase_regex = re.compile(r'[A-Z]')
    digit_regex = re.compile(r'\d')
    symbol_regex = re.compile(r'[\W_]')
    lowercase_regex = re.compile(r'[a-z]')
    # Vérifiez si le mot de passe répond aux critères
    length = length_regex.search(password)
    uppercase = uppercase_regex.search(password)
    digit = digit_regex.search(password)
    symbol = symbol_regex.search(password)
    lowercase = lowercase_regex.search(password)

    # Calculer le score et le message en fonction des critères remplis
    score = 0
    critere =''
    if length:
        score += 1
    else:
        critere += '               → be at least 8 characters.\n'

    if uppercase:
        score += 1
    else:
        critere += '            → contain capital letters.\n'
    
    if digit:
        score += 1
    else:
        critere += ' → contain numbers.\n'

    if symbol:
        score += 1
    else:
        critere += '→ contain symbols.\n'

    if lowercase:
        score += 1
    else:
        critere += '                   → contain lowercase letters.'

    return critere,score

def Critere(enent=None):
    password=mot_de_passe.get()
    c,score= check_password_strength(password)
    crit=Tk()
    crit.geometry('300x300+1050+300')
    crit.resizable(False,False)
    crit.title('Missing criteria')#F8F2ED
    crit.config(background='#F9F9F9') #pour determiner la color de background de la fenetre
    crit.iconbitmap('C:\\Users\\pc\\Desktop\\Python\\icon.ico')
    crit.attributes('-alpha',0.9)
    label = Label(crit, text='The password must :',bg='#F9F9F9',font=("Georgia",13))
    label.place(x=30,y=50)
    if(c==' → contain numbers.\n' or c=='→ contain symbols.\n' or c==' → contain numbers.\n'+'→ contain symbols.\n'):
        label = Label(crit,text=c,bg='#F9F9F9',font=("Impact",11))
        label.place(x=55,y=100)
    else:
        label = Label(crit,text=c,bg='#F9F9F9',font=("Impact",11))
        label.place(x=1,y=100)
    fermer=Button(crit, text="Close",bd=6,bg='#F9F9F9',fg='red',width='7',cursor='hand2',font=("Impact",10), command=crit.destroy)
    fermer.place(x=210 ,y=230)

def main(enent=None):
    password=mot_de_passe.get()
    critere,score = check_password_strength(password)
    force=Tk()
    force.geometry('450x450+700+200')
    force.resizable(False,False)
    force.title('Password Strength')#F8F2ED
    force.config(background='#F9F9F9') #pour determiner la color de background de la fenetre
    force.iconbitmap('C:\\Users\\pc\\Desktop\\Python\\icon.ico')
    force.attributes('-alpha',0.9)
    if score == 5:
        symbole=Label(force,text='✔',fg='green',bg='#F9F9F9',font=("Impact",50))
        symbole.place(x=195,y=100)
        text=Label(force,text='Your password is very strong!',bg='#F9F9F9',fg='green',font=("Impact",20))
        text.place(x=55,y=200)
        fermer=Button(force, text="Close",bd=6,bg='#F9F9F9',fg='red',width='25',cursor='hand2',font=("Impact",12), command=force.destroy)
        fermer.place(x=115 ,y=320)
        Cri='Very Strong'
        
    elif score >=4:
        symbole=Label(force,text='✍',bg='#F9F9F9',fg='#FF7415',font=("Impact",50))
        symbole.place(x=175,y=50)
        text=Label(force,text='Your password is strong.',bg='#F9F9F9',fg='#FF7415',font=("Impact",20))
        text.place(x=80,y=150)
        c=Button(force, text="see missing criteria",bd=6,width='25',cursor='hand2',font=("Impact",12),bg='#F9F9F9', fg='green', command=Critere)
        c.place(x=113 ,y=280)
        fermer=Button(force, text="Close",bd=6,bg='#F9F9F9',fg='red',width='25',cursor='hand2',font=("Impact",12), command=force.destroy)
        fermer.place(x=113 ,y=330)
        Cr='Strong'
        #cadre = LabelFrame(force, text="",bd=4)
        #label = Label(cadre, text=critere,bg='#F9F9F9',font=("Arial",10))
        #label.pack()
        #cadre.place(x=55,y=290)
        
    elif score == 3:
        symbole=Label(force,text='✍',bg='#F9F9F9',fg='#FF7415',font=("Impact",50))
        symbole.place(x=180,y=50)
        text=Label(force,text='Your password is correct.',fg='#FF7415',bg='#F9F9F9',font=("Impact",20))
        text.place(x=83,y=150)
        c=Button(force, text="see missing criteria",bd=6,width='25',cursor='hand2',font=("Impact",12),bg='#F9F9F9', fg='green', command=Critere)
        c.place(x=113 ,y=280)
        fermer=Button(force, text="Close",bd=6,bg='#F9F9F9',fg='red',width='25',cursor='hand2',font=("Impact",12), command=force.destroy)
        fermer.place(x=113 ,y=330)
        Cr='Correct'
        
    else:
        symbole=Label(force,text='✘',fg='red',bg='#F9F9F9',font=("Impact",50))
        symbole.place(x=180,y=80)
        text=Label(force,text= 'Your password is weak.',bg='#F9F9F9',fg='red',font=("Impact",20))
        text.place(x=90,y=180)
        #c=Label(force,text=critere,bg='#F9F9F9',fg='black',font=("Impact",12))
        #c.place(x=45,y=290)
        c=Button(force, text="see missing criteria",bd=6,width='25',cursor='hand2',font=("Impact",12),bg='#F9F9F9', fg='green', command=Critere)
        c.place(x=113 ,y=280)
        fermer=Button(force, text="Close",bd=6,bg='#F9F9F9',fg='red',width='25',cursor='hand2',font=("Impact",12), command=force.destroy)
        fermer.place(x=113 ,y=330)
        Cr='weak'

#fichier excel
Excel=Workbook()
fichier=Excel.active

fichier.title='Password History'
fichier['A1']='Full Name'
fichier['B1']='Email'
fichier['C1']='Password'
fichier['D1']='Type'

Excel.save('Password.xlsx')

def saveExcel(enent=None):
    name=En1.get()
    email=En2.get()
    password=mot_de_passe.get()
    c,score= check_password_strength(password)
    if score == 5:
        Cr='Very Strong'    
    elif score >=4:
        Cr='Strong'
    elif score == 3:
        Cr='Correct'   
    else:
        Cr='weak'

    excel=openpyxl.load_workbook('Password.xlsx')
    file=excel.active
    file.cell(column=1,row=file.max_row+1,value=name)
    file.cell(column=2,row=file.max_row,value=email)
    file.cell(column=3,row=file.max_row,value=password)
    file.cell(column=4,row=file.max_row,value=Cr)
    excel.save('Password.xlsx')

  
  
def main_valider(enent=None):
    main()
    saveExcel()


image = Image.open('C:\\Users\\pc\\Desktop\\Python\\image_login2.png')
photo = ImageTk.PhotoImage(image.resize((80, 80), Image.BICUBIC))
label_image = Label(fenetre, image=photo,bg='#002366')
label_image.place(x=202,y=50)

label1 = Label(fenetre, text="Name :",font=("Impact",11),bg='#002366',fg='#FF7415')
#label1.pack()
label1.place(x=120,y=180)
En1=Entry(fenetre,width='30',bd=4,relief=SUNKEN,font=("Arial", 11))
En1.place(x=125,y=210)

label2 = Label(fenetre, text="Email :",font=("Impact", 11),bg='#002366',fg='#FF7415')
#label2.pack()
label2.place(x=120,y=250)
En2=Entry(fenetre,width='30',bd=4,relief=SUNKEN,font=("Arial", 11))
En2.place(x=125,y=280)

label3 = Label(fenetre, text="Password :",font=("Impact", 11),bg='#002366',fg='#FF7415')
label3.place(x=121,y=320)    #label3.pack()
mot_de_passe = Entry(fenetre, show="* ",width='30',bd=4,relief=SUNKEN,font=("Arial", 11),justify=CENTER)
mot_de_passe.place(x=125,y=350)


test=Button(fenetre,text='Test',fg='#FF7415',bg='#002366',width='10',cursor='hand2',font=("Impact",12),bd=6,justify=LEFT,command=main_valider)
test.place(x=124 ,y=445)

reset=Button(fenetre,text='Reset',fg='#FF7415',bg='#002366',width='10',cursor='hand2',font=("Impact",12),bd=6,command=Reset,justify=RIGHT)
reset.place(x=275 ,y=445)
#fermer=Button(fenetre, text="Fermer", command=fenetre.quit)
#fermer.place(x=100 ,y=100)



fenetre.mainloop() #pour implementer tous les commends

